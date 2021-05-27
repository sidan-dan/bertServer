"""
this file aim to transfer the review_info to score info
@author:
create_time:
"""

import pymysql
import jieba
from openpyxl import load_workbook

class Scoring():
    def __init__(self):
        self.sen_dict = self.load_sen_dict()  # 加载情感词典
        print("成功加载情感词典")
        self.not_word_list = self.load_negative_data()  # 加载否定词列表
        print("成功加载否定词")
        self.degree_dic = self.load_degree_dict()  # 加载程度副词字典
        print("成功加载程度副词字典")

    def load_review_infos(self):
        db = pymysql.connect(host="localhost", user="root", password="hsk123456", port=3306, db="dianping",
                             charset="utf8")
        cursor = db.cursor()
        sql = "select id_review,shop_name,shop_id,shop_comment from dianping.shop_comment where id_review >825431"
        cursor.execute(sql)
        results = cursor.fetchall()  # 用于返回多条数据
        region_id = 'r35'
        sub_region_id = 'r1577'
        for i in results:#这里是每一条结果
            print(i)
            # 从结果的第一列开始用数字0开始排序，例如第六列，则其结果为i【5】
            key_id = i[0]
            shop_name = i[1]
            shop_id = i[2]
            review = i[3]
            word_info = self.cut_count_score(review)
            happy, reliable, surprise, praise, disappointed, anger = self.caculate_sen_value(word_info)
            score_sql = "insert into dianping.review_sen_score(id, region_id, sub_region_id, shop_name, shop_id, happy, praise, reliable, surprise, disppointed, anger,review) values ('{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}')".format(
                key_id, region_id, sub_region_id, shop_name, shop_id, happy, praise, reliable, surprise, disappointed,
                anger, review)
            try:
                if cursor.execute(score_sql):
                    print("Successful")

            except Exception as error:
                if str(error).find("SQL syntax") == -1:
                    print(error)
                print("failed")
                print(score_sql)
                db.rollback()
            db.commit()
        # 6.关闭查询
        cursor.close()
        # 关闭数据库
        db.close()
    def load_review(self,text):
        print(text)
        word_info = self.cut_count_score(text)
        print(word_info)
        happy, reliable, surprise, praise, disappointed, anger = self.caculate_sen_value(word_info)
        print('情绪值结果')
        print(happy, reliable, surprise, praise, disappointed, anger)
        value = 0
        if disappointed + anger + happy + surprise + praise / 5 + reliable / 20 < 0:
            if (disappointed + anger > - 10):
                value = 45 + happy + surprise + disappointed / 3 + anger / 3
            elif (disappointed + anger < -40) and (surprise + happy < 20):
                value = 30 + happy / 20 + surprise / 20 + disappointed / 50 + anger / 50
            else:
                value = 50 + happy / 5 + surprise / 5 + reliable / 20 + praise / 15 + disappointed / 3 + anger / 2
                value = value + (happy / 10) * 5 + (anger / 10) * 5
            print(value)
            return happy, reliable, surprise, praise, disappointed, anger, value

        if disappointed + anger + happy + surprise + praise / 5 + reliable / 20 > 0:
            if (happy + surprise < 10):
                value = 80 + happy + surprise + disappointed / 3 + anger / 3
            elif (happy + surprise > 40) and (disappointed + anger > -20):
                value = 90 + happy / 50 + surprise / 50 + disappointed / 20 + anger / 20
            else:
                value = 65 + happy / 2 + surprise / 2 + reliable / 20 + praise / 15 + disappointed / 10 + anger / 10
                value = value + (happy / 10) * 5 + (anger / 10) * 5
            print(value)
            return happy, reliable, surprise, praise, disappointed, anger, value
        if disappointed + anger + happy + surprise + praise / 5 + reliable / 20 == 0:
            value = 50 + happy / 10 + surprise / 10 + reliable / 20 + praise / 15 + disappointed / 10 + anger / 10
            value = value + (happy / 10) * 5 + (anger / 10) * 5
            print(value)
            return happy, reliable, surprise, praise, disappointed, anger, value
        print(value)
        return happy, reliable, surprise, praise, disappointed, anger,value

        # #多文本测试
        # reviews = ['当自我放纵成为了一种习惯，我会不自觉地为自己的放纵找各种各样的借口，借此来证明放纵是合理的。但是事实总是会清醒地告诉你：你一直在浪费时间，在浪费你美好的青春，纵欲只能为你带来低级的快乐，并且在快乐过后，会陷入无尽的自责深渊，纵欲带来的负面影响是你短时间自虐式努力挽回不了的。',
        #            '爱情，理想？这些所谓带着世间美好期待却又裹挟着一部分黑暗、凋败、欺骗的事物，是危险又迷人的。为何出走半生好像早已老道的人们会频频陷入愚蠢，一头栽进这个浩荡没定数的漩涡，大抵都是心里那份痴念吧。',
        #            '今天阳光不大，云只有薄薄的的几朵，天空格外的蓝，让人觉得这并不像刚刚下过雨。西边不知何时出现了一道彩虹。它很完整，真的就像一道桥，想人踏上去看看上面的美好。不像自然所谓，就好像有人乘着电梯在上面建造而成的。因为距离十分的遥远，显得十分的渺小，想伸出手去抓，然后碰到手心里，可惜根本抓不到。眼前忽然间出现了一堆黑色的乱线，这到底是电线，还是毛线。一直挡在眼前，无论怎么变换位置都再也看不到彩虹的全貌了。这究竟是为何，这究竟是为何。拿着相机呆呆的站着。叮叮，手机铃响了，彩虹不见了，一切只是一场梦。下了床，和往常一样到阳台拿衣服。下意识向窗外看，没有蓝天，只有电线杆。',
        #            '从小学一年级到步入大学从6岁的开始到现在19岁的成年不知不觉已13年了你占据了我生命的三分之二见过你笑见过你哭见过你的点点滴滴也记得我们的狼狈不堪深夜畅谈肆无忌惮很感谢你出现在我的生命里陪伴我成长好像也说不出什么矫情有深度的话在未来的时光里愿你被温柔对待愿你开心愿你幸福愿有对你很好的人出现早日脱单愿我们能一直相伴无论何时 我一直在',
        #            '今天见到了小白真的很开心',
        #            '哎，什么时候才能又见到她啊，难受']
        # for i in reviews:
        #     word_info = self.cut_count_score(i)
        #     happy, reliable, surprise, praise, disappointed, anger = self.caculate_sen_value(word_info)
        #     print('情绪值结果')
        #     print( happy, reliable, surprise, praise, disappointed, anger )
        #     value = 0
        #     # value = 50 + happy + reliable + surprise + praise + disappointed + anger
        #     if disappointed+anger+happy+surprise+praise/5+ reliable/20<0:
        #         if (disappointed + anger > - 10):
        #             value = 45 + happy + surprise + disappointed / 3 + anger / 3
        #         elif (disappointed + anger < -40) and (surprise + happy < 20):
        #             value = 30 + happy / 20 + surprise / 20 + disappointed / 50 + anger / 50
        #         else:
        #             value = 50 + happy/5 + surprise/5 +reliable/20 + praise/15 + disappointed/3 +anger/2
        #             value = value + (happy / 10) * 5 + (anger / 10) * 5
        #     if disappointed + anger + happy + surprise+ praise/5 + reliable/20> 0:
        #         if(happy+surprise<10):
        #             value = 80 + happy +surprise + disappointed / 3 + anger / 3
        #         elif(happy+surprise>40)and(disappointed+anger>-20):
        #             value = 90 +happy/50 + surprise /50 +disappointed/20 +anger/20
        #         else:
        #             value = 65 + happy / 2 + surprise / 2 + reliable / 20 + praise / 15 + disappointed / 10 + anger / 10
        #             value = value + (happy / 10) * 5 + (anger / 10) * 5
        #     if disappointed + anger + happy + surprise +praise/5 + reliable/20== 0:
        #         value = 50 + happy / 10 + surprise / 10 + reliable / 20 + praise / 15 + disappointed / 10 + anger / 10
        #         value = value + (happy / 10) * 5 + (anger / 10) * 5
        #     print(value)
        # return happy, reliable, surprise, praise, disappointed, anger,value
    def need_data(self):
        import openpyxl
        filename = "/Users/sikehu/PycharmProjects/pythonProject/caculate_review_emotion_score/origin_modify/need_data.xlsx"
        wb = load_workbook(filename)
        a_sheet = wb['Sheet1']
        rows = a_sheet.max_row  # 得到行数
        ids = []  # 用于存储程度副词的词语
        texts = []  # 用于存储程度副词的分数
        for i in range(1, rows + 1):
            n1 = a_sheet.cell(row=i, column=1).value
            n2 = a_sheet.cell(row=i, column=2).value
            ids.append(n1)
            texts.append(n2)
        workbook = openpyxl.Workbook()
        worksheet = workbook.create_sheet(title="Sheet1", index=0)
        row = 1
        for text in texts:
            print("-" * 100)
            print(text)
            word_info = self.cut_count_score(text)
            happy, reliable, surprise, praise, disappointed, anger = self.caculate_sen_value(word_info)
            print(happy, reliable, surprise, praise, disappointed, anger)
    # def store_to_excel(self):
    #     Key_id, Shop_name, Shop_id, Review = self.load_review_infos()
    #     mysql = pymysql.connect(host="localhost", user="root", password="hsk123456", port=3306, db="dianping",
    #                          charset="utf8")
    #     cursor1 = mysql.cursor()
    #     num = 0
    #     number = len(Key_id)
    #     region_id = 'r35'
    #     sub_region_id = 'r1577'
    #     while (num < number):
    #         key_id = Key_id[num]
    #         shop_name = Shop_name[num]
    #         shop_id = Shop_id[num]
    #         review = Review[num]
    #         word_info = self.cut_count_score(review)
    #         happy, reliable, surprise, praise, disappointed, anger = self.caculate_sen_value(word_info)
    #         score_sql = "insert into dianping.review_sen_score(id, region_id, sub_region_id, shop_name, shop_id, happy, praise, reliable, surprise, disppointed, anger,review) values ('{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}')".format(
    #             key_id, region_id, sub_region_id, shop_name, shop_id, happy, praise, reliable, surprise, disappointed,
    #             anger, review)
    #         try:
    #             if cursor1.execute(score_sql):
    #                 print("Successful")
    #                 print(score_sql)
    #
    #         except Exception as error:
    #             if str(error).find("SQL syntax") == -1:
    #                 print(error)
    #             print("failed")
    #             print(score_sql)
    #         num += 1
    #     # 6.关闭查询
    #     cursor1.close()
    #     # 关闭数据库
    #     mysql.close()

    def stopwordslist(self):
        """
        this file aim to load stopwords
        :return:
        """
        stopwords = [line.strip() for line in open('stopword.txt', encoding='UTF-8').readlines()]
        return stopwords

    def load_self_words(self):
        """
        this file aim to load self defination word
        :return:
        """
        # 加载自定义词汇
        jieba.add_word('宽巷子', freq=20000)
        jieba.add_word('子非', freq=20000)
        jieba.add_word('麻辣', freq=20000)
        jieba.add_word('私房菜', freq=20000)
        jieba.add_word('毛笔酥', freq=20000)
        jieba.add_word('定制款', freq=20000)
        jieba.add_word('行酒令', freq=20000)
        jieba.add_word('摇色子', freq=20000)
        jieba.add_word('米翘', freq=20000)
        jieba.add_word('宽窄巷子', freq=20000)
        jieba.add_word('播音腔', freq=20000)
        jieba.add_word('不同', freq=20000)
        jieba.add_word('刚刚熟', freq=20000)
        jieba.add_word('外酥里嫩', freq=20000)
        jieba.add_word('一级棒', freq=20000)
        jieba.add_word('店面设计', freq=20000)
        jieba.add_word('动筷子', freq=20000)
        jieba.add_word('川渝', freq=20000)
        jieba.add_word('工匠精神', freq=20000)
        jieba.add_word('毫不夸张', freq=20000)
        jieba.add_word('仪式感', freq=20000)
        jieba.add_word('商务宴请', freq=20000)
        jieba.add_word('夫妻肺片', freq=20000)
        jieba.add_word('没得挑', freq=20000)
        jieba.add_word('大众点评', freq=20000)
        jieba.add_word('一脸懵逼', freq=20000)
        jieba.add_word('没什么特别', freq=20000)
        jieba.add_word('创意菜', freq=20000)
        jieba.add_word('米二', freq=20000)
        jieba.add_word('意境菜', freq=20000)
        jieba.add_word('早有耳闻', freq=20000)
        jieba.add_word('老宅子', freq=20000)
        jieba.add_word('一步一景', freq=20000)
        jieba.add_word('嚼不动', freq=20000)
        jieba.add_word('加分项目', freq=20000)
        jieba.add_word('黑珍珠餐厅', freq=20000)
        jieba.add_word('么么哒', freq=20000)
        jieba.add_word('电话预约', freq=20000)
        jieba.add_word('商务接待', freq=20000)
        jieba.add_word('私房菜馆', freq=20000)
        jieba.add_word('文艺范', freq=20000)
        jieba.add_word('富贵花开', freq=20000)
        jieba.add_word('金熊猫餐厅', freq=20000)
        jieba.add_word('Q弹', freq=20000)
        jieba.add_word('大v', freq=20000)
        jieba.add_word('兰亭序', freq=20000)
        jieba.add_word('超级贵', freq=20000)
        jieba.add_word('很辣', freq=20000)
        jieba.add_word('源远文化', freq=20000)
        jieba.add_word('莲蓉千层酥', freq=20000)
        jieba.add_word('宰游客', freq=20000)
        jieba.add_word('年代感', freq=20000)
        jieba.add_word('打飞的', freq=20000)
        jieba.add_word('鸡蛋里挑骨头', freq=20000)
        jieba.add_word('金熊猫', freq=20000)
        jieba.add_word('出挑', freq=20000)
        jieba.add_word('清洁阿姨', freq=20000)
        jieba.add_word('烘烘', freq=20000)
        jieba.add_word('良心商家', freq=20000)
        jieba.add_word('寬窄巷子', freq=20000)
        jieba.add_word('花钱买教训', freq=20000)
        jieba.add_word('夫妻肺片', freq=20000)
        jieba.add_word('别有一番滋味', freq=20000)
        jieba.add_word('宫保鸡丁', freq=20000)
        jieba.add_word('菌菇汤', freq=20000)
        jieba.add_word('彭子渝', freq=20000)
        jieba.add_word('体验感', freq=20000)
        jieba.add_word('慢生活', freq=20000)
        jieba.add_word('地铁口', freq=20000)
        jieba.add_word('个人喜好', freq=20000)
        jieba.add_word('萌萌哒', freq=20000)
        jieba.add_word('兰亭叙', freq=20000)
        jieba.add_word('太古里', freq=20000)
        jieba.add_word('中规中矩', freq=20000)
        jieba.add_word('中矩中规', freq=20000)
        jieba.add_word('刷锅水', freq=20000)
        jieba.add_word('热门菜套路', freq=20000)
        jieba.add_word('老天不作美', freq=20000)
        jieba.add_word('不得不说', freq=20000)
        jieba.add_word('宝宝椅', freq=20000)
        jieba.add_word('兜兜转转', freq=20000)
        jieba.add_word('蜀江春', freq=20000)
        jieba.add_word('有颜有味', freq=20000)
        jieba.add_word('中医大附院', freq=20000)
        jieba.add_word('现杀现做', freq=20000)
        jieba.add_word('玉米汁', freq=20000)
        jieba.add_word('清江东路店', freq=20000)
        jieba.add_word('科华路店', freq=20000)
        jieba.add_word('自贡菜', freq=20000)
        jieba.add_word('川菜馆子', freq=20000)
        jieba.add_word('仔姜鸭', freq=20000)
        jieba.add_word('省医院', freq=20000)
        jieba.add_word('热毛巾', freq=20000)
        jieba.add_word('双人餐', freq=20000)
        jieba.add_word('棒棒哒', freq=20000)
        jieba.add_word('双人套餐', freq=20000)
        jieba.add_word('温哥华广场', freq=20000)
        jieba.add_word('青羊小区', freq=20000)
        jieba.add_word('点赞', freq=20000)
        jieba.add_word('倍加', freq=20000)
        jieba.add_word('必定', freq=20000)
        jieba.add_word('必须', freq=20000)
        jieba.add_word('并非不', freq=20000)
        jieba.add_word('不得不', freq=20000)
        jieba.add_word('不妨', freq=20000)
        jieba.add_word('常常', freq=20000)
        jieba.add_word('超', freq=20000)
        jieba.add_word('超级', freq=20000)
        jieba.add_word('处处', freq=20000)
        jieba.add_word('从来', freq=20000)
        jieba.add_word('大', freq=20000)
        jieba.add_word('大概', freq=20000)
        jieba.add_word('大约', freq=20000)
        jieba.add_word('单', freq=20000)
        jieba.add_word('单单', freq=20000)
        jieba.add_word('的确', freq=20000)
        jieba.add_word('都', freq=20000)
        jieba.add_word('顿时', freq=20000)
        jieba.add_word('多', freq=20000)
        jieba.add_word('非常', freq=20000)
        jieba.add_word('分外', freq=20000)
        jieba.add_word('赶紧', freq=20000)
        jieba.add_word('感觉', freq=20000)
        jieba.add_word('格外', freq=20000)
        jieba.add_word('更', freq=20000)
        jieba.add_word('更加', freq=20000)
        jieba.add_word('更其', freq=20000)
        jieba.add_word('更为', freq=20000)
        jieba.add_word('共', freq=20000)
        jieba.add_word('过于', freq=20000)
        jieba.add_word('还', freq=20000)
        jieba.add_word('还是', freq=20000)
        jieba.add_word('何等', freq=20000)
        jieba.add_word('很', freq=20000)
        jieba.add_word('极', freq=20000)
        jieba.add_word('极度', freq=20000)
        jieba.add_word('极端', freq=20000)
        jieba.add_word('极力', freq=20000)
        jieba.add_word('极其', freq=20000)
        jieba.add_word('极为', freq=20000)
        jieba.add_word('几乎', freq=20000)
        jieba.add_word('简直', freq=20000)
        jieba.add_word('渐渐', freq=20000)
        jieba.add_word('将要', freq=20000)
        jieba.add_word('较', freq=20000)
        jieba.add_word('较比', freq=20000)
        jieba.add_word('较为', freq=20000)
        jieba.add_word('仅仅', freq=20000)
        jieba.add_word('尽', freq=20000)
        jieba.add_word('进一步', freq=20000)
        jieba.add_word('净', freq=20000)
        jieba.add_word('居然', freq=20000)
        jieba.add_word('决非不', freq=20000)
        jieba.add_word('绝顶', freq=20000)
        jieba.add_word('绝对', freq=20000)
        jieba.add_word('立刻', freq=20000)
        jieba.add_word('连忙', freq=20000)
        jieba.add_word('略微', freq=20000)
        jieba.add_word('蛮', freq=20000)
        jieba.add_word('偶尔', freq=20000)
        jieba.add_word('偏偏', freq=20000)
        jieba.add_word('颇', freq=20000)
        jieba.add_word('恰恰', freq=20000)
        jieba.add_word('强', freq=20000)
        jieba.add_word('全', freq=20000)
        jieba.add_word('稍', freq=20000)
        jieba.add_word('稍微', freq=20000)
        jieba.add_word('甚为', freq=20000)
        jieba.add_word('十分', freq=20000)
        jieba.add_word('十足', freq=20000)
        jieba.add_word('时常', freq=20000)
        jieba.add_word('始终', freq=20000)
        jieba.add_word('太', freq=20000)
        jieba.add_word('忒', freq=20000)
        jieba.add_word('特', freq=20000)
        jieba.add_word('特别', freq=20000)
        jieba.add_word('挺', freq=20000)
        jieba.add_word('统统', freq=20000)
        jieba.add_word('完全', freq=20000)
        jieba.add_word('往往', freq=20000)
        jieba.add_word('未', freq=20000)
        jieba.add_word('未免', freq=20000)
        jieba.add_word('无比', freq=20000)
        jieba.add_word('无不', freq=20000)
        jieba.add_word('相当', freq=20000)
        jieba.add_word('幸而', freq=20000)
        jieba.add_word('幸亏', freq=20000)
        jieba.add_word('也许', freq=20000)
        jieba.add_word('一点都', freq=20000)
        jieba.add_word('一点也', freq=20000)
        jieba.add_word('一概', freq=20000)
        jieba.add_word('一律', freq=20000)
        jieba.add_word('一齐', freq=20000)
        jieba.add_word('一向', freq=20000)
        jieba.add_word('尤其', freq=20000)
        jieba.add_word('尤为', freq=20000)
        jieba.add_word('有点儿', freq=20000)
        jieba.add_word('愈', freq=20000)
        jieba.add_word('愈加', freq=20000)
        jieba.add_word('愈来愈', freq=20000)
        jieba.add_word('愈益', freq=20000)
        jieba.add_word('远远', freq=20000)
        jieba.add_word('越', freq=20000)
        jieba.add_word('越发', freq=20000)
        jieba.add_word('越加', freq=20000)
        jieba.add_word('越来越', freq=20000)
        jieba.add_word('越是', freq=20000)
        jieba.add_word('再', freq=20000)
        jieba.add_word('再三', freq=20000)
        jieba.add_word('早已', freq=20000)
        jieba.add_word('这般', freq=20000)
        jieba.add_word('这样', freq=20000)
        jieba.add_word('只', freq=20000)
        jieba.add_word('只好', freq=20000)
        jieba.add_word('终于', freq=20000)
        jieba.add_word('准', freq=20000)
        jieba.add_word('总', freq=20000)
        jieba.add_word('总共', freq=20000)
        jieba.add_word('总是', freq=20000)
        jieba.add_word('足', freq=20000)
        jieba.add_word('足足', freq=20000)
        jieba.add_word('最', freq=20000)
        jieba.add_word('最为', freq=20000)
        jieba.add_word('至极', freq=20000)
        jieba.add_word('之至', freq=20000)
        jieba.add_word('之极', freq=20000)
        jieba.add_word('贼', freq=20000)
        jieba.add_word('异常', freq=20000)
        jieba.add_word('万万', freq=20000)
        jieba.add_word('万般', freq=20000)
        jieba.add_word('滔天', freq=20000)
        jieba.add_word('十二分', freq=20000)
        jieba.add_word('莫大', freq=20000)
        jieba.add_word('奇', freq=20000)
        jieba.add_word('满心', freq=20000)
        jieba.add_word('刻骨', freq=20000)
        jieba.add_word('不堪', freq=20000)
        jieba.add_word('百分之百', freq=20000)
        jieba.add_word('备至', freq=20000)
        jieba.add_word('不得了', freq=20000)
        jieba.add_word('不可开交', freq=20000)
        jieba.add_word('不亦乐乎', freq=20000)
        jieba.add_word('不折不扣', freq=20000)
        jieba.add_word('彻头彻尾', freq=20000)
        jieba.add_word('充分', freq=20000)
        jieba.add_word('到头', freq=20000)
        jieba.add_word('地地道道', freq=20000)
        jieba.add_word('截然', freq=20000)
        jieba.add_word('惊人地', freq=20000)
        jieba.add_word('绝对化', freq=20000)
        jieba.add_word('酷', freq=20000)
        jieba.add_word('满贯', freq=20000)
        jieba.add_word('入骨', freq=20000)
        jieba.add_word('死', freq=20000)
        jieba.add_word('痛', freq=20000)
        jieba.add_word('透', freq=20000)
        jieba.add_word('完完全全', freq=20000)
        jieba.add_word('万', freq=20000)
        jieba.add_word('万分', freq=20000)
        jieba.add_word('无度', freq=20000)
        jieba.add_word('无可估量', freq=20000)
        jieba.add_word('无以复加', freq=20000)
        jieba.add_word('无以伦比', freq=20000)
        jieba.add_word('要命', freq=20000)
        jieba.add_word('要死', freq=20000)
        jieba.add_word('已极', freq=20000)
        jieba.add_word('已甚', freq=20000)
        jieba.add_word('逾常', freq=20000)
        jieba.add_word('卓绝', freq=20000)
        jieba.add_word('佼佼', freq=20000)
        jieba.add_word('不少', freq=20000)
        jieba.add_word('不胜', freq=20000)
        jieba.add_word('惨', freq=20000)
        jieba.add_word('沉', freq=20000)
        jieba.add_word('沉沉', freq=20000)
        jieba.add_word('出奇', freq=20000)
        jieba.add_word('大为', freq=20000)
        jieba.add_word('多多', freq=20000)
        jieba.add_word('多加', freq=20000)
        jieba.add_word('多么', freq=20000)
        jieba.add_word('够瞧的', freq=20000)
        jieba.add_word('够呛', freq=20000)
        # jieba.add_word('好', freq=20000)
        jieba.add_word('好不', freq=20000)
        jieba.add_word('很是', freq=20000)
        jieba.add_word('坏', freq=20000)
        jieba.add_word('可', freq=20000)
        jieba.add_word('老', freq=20000)
        jieba.add_word('老大', freq=20000)
        jieba.add_word('良', freq=20000)
        jieba.add_word('颇为', freq=20000)
        jieba.add_word('甚', freq=20000)
        jieba.add_word('实在', freq=20000)
        jieba.add_word('太甚', freq=20000)
        jieba.add_word('远', freq=20000)
        jieba.add_word('着实', freq=20000)
        jieba.add_word('大不了', freq=20000)
        jieba.add_word('更进一步', freq=20000)
        jieba.add_word('还要', freq=20000)
        jieba.add_word('那般', freq=20000)
        jieba.add_word('那么', freq=20000)
        jieba.add_word('那样', freq=20000)
        jieba.add_word('如斯', freq=20000)
        jieba.add_word('尤甚', freq=20000)
        jieba.add_word('怪', freq=20000)
        jieba.add_word('好生', freq=20000)
        jieba.add_word('或多或少', freq=20000)
        jieba.add_word('略', freq=20000)
        jieba.add_word('略加', freq=20000)
        jieba.add_word('略略', freq=20000)
        jieba.add_word('略为', freq=20000)
        jieba.add_word('稍稍', freq=20000)
        jieba.add_word('稍为', freq=20000)
        jieba.add_word('稍许', freq=20000)
        jieba.add_word('些微', freq=20000)
        jieba.add_word('一点', freq=20000)
        jieba.add_word('一点儿', freq=20000)
        jieba.add_word('一些', freq=20000)
        jieba.add_word('不大', freq=20000)
        jieba.add_word('半点', freq=20000)
        jieba.add_word('不定点儿', freq=20000)
        jieba.add_word('不甚', freq=20000)
        jieba.add_word('不怎么', freq=20000)
        jieba.add_word('没怎么', freq=20000)
        jieba.add_word('轻度', freq=20000)
        jieba.add_word('弱', freq=20000)
        jieba.add_word('丝毫', freq=20000)
        jieba.add_word('相对', freq=20000)
        jieba.add_word('不为过', freq=20000)
        jieba.add_word('超额', freq=20000)
        jieba.add_word('过度', freq=20000)
        jieba.add_word('过分', freq=20000)
        jieba.add_word('过火', freq=20000)
        jieba.add_word('过劲', freq=20000)
        jieba.add_word('过猛', freq=20000)
        jieba.add_word('过热', freq=20000)
        jieba.add_word('过甚', freq=20000)
        jieba.add_word('过头', freq=20000)
        jieba.add_word('何止', freq=20000)
        jieba.add_word('偏', freq=20000)
        jieba.add_word('何况', freq=20000)
        jieba.add_word('况且', freq=20000)
        jieba.add_word('不但', freq=20000)
        jieba.add_word('不仅', freq=20000)
        jieba.add_word('连', freq=20000)
        jieba.add_word('不但不', freq=20000)
        jieba.add_word('甚至', freq=20000)
        jieba.add_word('原来', freq=20000)
        jieba.add_word('由于', freq=20000)
        jieba.add_word('以便', freq=20000)
        jieba.add_word('所以', freq=20000)
        jieba.add_word('以致', freq=20000)
        jieba.add_word('因此', freq=20000)
        jieba.add_word('既然', freq=20000)
        jieba.add_word('可见', freq=20000)
        jieba.add_word('就', freq=20000)
        jieba.add_word('则', freq=20000)
        jieba.add_word('但', freq=20000)
        jieba.add_word('却', freq=20000)
        jieba.add_word('然而', freq=20000)
        jieba.add_word('但是', freq=20000)
        jieba.add_word('只不过', freq=20000)
        jieba.add_word('可是', freq=20000)
        jieba.add_word('尽管', freq=20000)
        jieba.add_word('不过', freq=20000)
        jieba.add_word('即使', freq=20000)
        jieba.add_word('若', freq=20000)
        jieba.add_word('假如', freq=20000)
        jieba.add_word('如果', freq=20000)
        jieba.add_word('假使', freq=20000)
        jieba.add_word('纵是', freq=20000)
        jieba.add_word('纵然', freq=20000)

    def cut_count_score(self,review):
        """
        this file aim to get the emotion score of the review
        :param review:
        :return:
        """
        word_info = []
        # 创建一个停用词列表
        stopwords = self.stopwordslist()
        #加载自定义词汇
        self.load_self_words()
        sentence = str(review)
        if sentence:
            sentence = sentence.replace('\u202a', '')
            sentence = sentence.replace('\u202b', '')
            sentence = sentence.replace('\u202c', '')
            sentence = sentence.replace('\u202d', '')
            sentence = sentence.replace('\u202e', '')
            sentence = sentence.replace('\u202f', '')
            sentence = sentence.replace(' ', '')

            sentence_depart = jieba.lcut(sentence.strip())
            word_loc = 0
            for word in sentence_depart:
                if word not in stopwords:
                    word_info.append((word,word_loc))
                    word_loc += 1
        return word_info

    def load_sen_dict(self):  # 智力	adj	PH	5  只重复了这一个，把这个单独拎出来
        """
        # 将词语作为主键加载情感词典
        :return:
        """
        wb = load_workbook("emotionwords(1).xlsx")
        a_sheet = wb['Sheet1']
        rows = a_sheet.max_row  # 得到行数
        # print("rows:",rows)
        word = []  # 用于存储词语
        part_of_speech = []  # 用于存储词性种类
        s_classify = []  # 用于存储词语情感分类
        s_value = []  # 用于存储词语情感值
        for i in range(2, rows + 1):
            n1 = a_sheet.cell(row=i, column=2).value
            n2 = a_sheet.cell(row=i, column=3).value
            n3 = a_sheet.cell(row=i, column=4).value
            n4 = a_sheet.cell(row=i, column=5).value
            word.append(n1)
            part_of_speech.append(n2)
            s_classify.append(n3)
            s_value.append(n4)
        emotion_dict = {}
        i_degree = 0
        while i_degree < (rows-1):
            emotion_dict[word[i_degree]] = [part_of_speech[i_degree], s_classify[i_degree], s_value[i_degree]]
            i_degree += 1
        # #读取时：
        # for key,value in emotion_dict.items():
        #     word = key[0]
        #     speech = key[1]
        #     classify = value[0]
        #     s_va = value[1]
        return emotion_dict

    def load_negative_data(self):
        """
        加载否定词
        :return:
        """
        wb = load_workbook("not_words.xlsx")
        a_sheet = wb['Sheet1']
        rows = a_sheet.max_row  # 得到行数
        # print("max rows:",rows)
        not_words = []
        for i in range(1, rows + 1):
            n1 = a_sheet.cell(row=i, column=1).value
            not_words.append(n1)
        # print("data load successful!!!")  #在这里我读取了否定词
        return not_words  # 返回了否定词列表

    def load_degree_dict(self):
        """
        # 加载程度副词
        :return:
        """
        filename = "Adverbs_of_degree.xlsx"
        wb = load_workbook(filename)
        a_sheet = wb['Sheet1']
        rows = a_sheet.max_row  # 得到行数
        degree_words = []  # 用于存储程度副词的词语
        degree_scores = []  # 用于存储程度副词的分数
        for i in range(1, rows + 1):
            n1 = a_sheet.cell(row=i, column=1).value
            n2 = a_sheet.cell(row=i, column=2).value
            degree_words.append(n1)
            degree_scores.append(n2)
        degree_dict = {}
        i_degree = 0
        while i_degree < rows:
            degree_dict[str(degree_words[i_degree])] = degree_scores[i_degree]
            i_degree += 1
        return degree_dict

    def caculate_sen_value(self,info):
        """
        传入句子的分词，计算句子的情感值
        :param info:
        :return:
        """
        sen_word = {}
        not_word = {}
        degree_word = {}
        """计算得分"""
        # 权重初始化为1
        W = 1
        score = 0
        loc = 0
        # 初始化各个情感值
        happy = 0
        reliable = 0
        praise = 0
        surprise = 0
        disappointed = 0
        anger = 0
        # 该for循环将词语分为了情感词、否定词和程度副词，为计算情感做铺垫
        for word_info in info:
            word = word_info[0]  # 词语 eg:去年
            # loc = word_info[1]   # 词性 eg:prep
            if word in self.sen_dict.keys() and word not in self.not_word_list and word not in self.degree_dic.keys():
                sen_word[loc] = [self.sen_dict[word][1], self.sen_dict[word][2]]  # 情感词，主键为位置，值为词语的情感属性和情感值
                # sen_word[loc] = [sen_dict[word][1],sen_dict[word][2],word]   #情感词，主键为位置，值为词语的情感属性和情感值
                print("sen-word:", word,self.sen_dict[word][2])
                loc += 1
            elif word in self.not_word_list and word not in self.degree_dic.keys():
                # 分词结果中在否定词列表中的词
                # not_word[loc] = [-1,word]
                not_word[loc] = -1
                print("not-word:", word,-1)
                loc += 1
            elif word in self.degree_dic.keys():
                # 分词结果中在程度副词中的词
                degree_word[loc] = self.degree_dic[word]
                # degree_word[loc] = [degree_dic[word],word]
                print("degree-word:", word,self.degree_dic[word])
                loc += 1
            # print(num)
        print("sen_dict",sen_word)
        print("not_word",not_word)
        print("degree_word",degree_word)
            # print(loc)  #此处loc就是senword，notword和degreeword一共的个数
        for i_loc in range(0, loc):  # 该for循环用于将否定词和程度副词赋值给修饰的情感词，i_loc用于控制位置
            # 若是否定词
            if i_loc in not_word.keys():
                not_value = not_word[i_loc]
                if (i_loc + 1) in not_word.keys():
                    not_word[(i_loc + 1)] = float(not_word[(i_loc + 1)]) * float(not_value)
                elif (i_loc + 1) in degree_word.keys():
                    degree_word[(i_loc + 1)] = float(degree_word[(i_loc + 1)]) * float(not_value)
                elif (i_loc + 1) in sen_word.keys():
                    sen_word[(i_loc + 1)][1] = float(sen_word[(i_loc + 1)][1]) * float(not_value)
            # 若是程度副词
            elif i_loc in degree_word.keys():
                degree_value = degree_word[i_loc]
                if (i_loc + 1) in not_word.keys():
                    not_word[(i_loc + 1)] = float(not_word[(i_loc + 1)]) * float(degree_value)
                elif (i_loc + 1) in degree_word.keys():
                    degree_word[(i_loc + 1)] = float(degree_word[(i_loc + 1)]) * float(degree_value)
                elif (i_loc + 1) in sen_word.keys():
                    sen_word[(i_loc + 1)][1] = float(sen_word[(i_loc + 1)][1]) * float(degree_value)
        # print(sen_word)
        for loc_i in range(0, loc):  # 该for循环用于将情感值分类并且计算
            if loc_i in sen_word.keys():
                emotion = sen_word[loc_i][0]  # 词语的情感分类
                sen_value = sen_word[loc_i][1]  # 词语的情感值
                # 用于判断该词属于哪一个情感
                if emotion in ['PA']:
                    happy += sen_value
                if emotion in ['PE', 'PG']:
                    reliable += sen_value
                if emotion in ['PD', 'PH', 'PB', 'PK', 'PF']:
                    praise += sen_value
                    # print("praise:",praise)
                    # print("sen_value:",sen_value)
                if emotion in ['PC']:
                    surprise += sen_value
                if emotion in ['NA', 'NE', 'ND', 'NN']:
                    anger += sen_value
                if emotion in ['NB', 'NJ']:
                    disappointed += sen_value
            # 计算这一个id的总情感
        anger = anger * (-1)
        disappointed = disappointed * (-1)
        return happy,reliable,surprise,praise,disappointed,anger


if __name__ == '__main__':
    obj = Scoring()
    obj.load_review()
    # load_yunyun()
    # obj.need_data()

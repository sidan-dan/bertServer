# pycharm项目使用anaconda环境tensorflow
# bert启动使用anaconda环境py3.5
import json
import os
import datetime

import random
from imp import reload

import numpy
import xlrd
import sys

from flask import Flask, request, make_response, jsonify, render_template, send_from_directory, Response
from openpyxl import load_workbook

from xlutils.copy import copy
from numpy import argmax

from bad import Clustering
from caculate_review_emotion_score import Scoring
from nlp_classify import Classify

app = Flask(__name__)

UPLOAD_FOLDER = 'upload'
DOWNLOAD_FOLDER = 'download'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['DOWNLOAD_FOLDER'] = DOWNLOAD_FOLDER
basedir = os.path.abspath(os.path.dirname(__file__))
ALLOWED_EXTENSIONS = set(['xls', 'xlsx'])
obj = Scoring()

class Pic_str:
    def create_uuid(self): #生成唯一的源文件的名称字符串，防止图片显示时的重名问题
        nowTime = datetime.datetime.now().strftime("%Y%m%d%H%M%S");  # 生成当前时间
        randomNum = random.randint(0, 100);  # 生成的随机整数n，其中0<=n<=100
        if randomNum <= 10:
            randomNum = str(0) + str(randomNum);
        uniqueNum = str(nowTime) + str(randomNum);
        return uniqueNum;

@app.route('/analyseData', methods=['GET','POST'], strict_slashes=False)
def hello_world():
    print("进入测试")
    file_dir = os.path.join(basedir, app.config['UPLOAD_FOLDER'])
    if not os.path.exists(file_dir):
        os.makedirs(file_dir)
    download_dir = os.path.join(basedir, app.config['DOWNLOAD_FOLDER'])
    if not os.path.exists(download_dir):
        os.makedirs(download_dir)
    # 储存源文件
    f = request.files['wximage']
    if f and allowed_file(f.filename):
        fname = f.filename
        print(fname)
        ext = fname.rsplit('.', 1)[1]
        new_filename = Pic_str().create_uuid() + '.' + ext
        print(new_filename)
        f.save(os.path.join(file_dir, new_filename))

    # 加载execl数据
    excel_path = os.path.join(file_dir, new_filename)  # Excel文件路径
    # excel_path = "C:/Users/司丹/Desktop/test.xls"
    print('Excel文件的路径：' + excel_path)
    excel_file = xlrd.open_workbook(excel_path)  # 打开Excel文件
    table = excel_file.sheets()[0]  # 通过索引打开
    print('已经打开的工作簿的名字：' + table.name)
    print('**********开始读取Excel单元格的内容**********')
    all_content = []
    for i in range(table.nrows):
        cell_value = table.cell(i, 0).value  # 获取单元格内容
        all_content.append(cell_value)
        # Logger().info('[' + ', '.join("'" + str(element) + "'" for element in row_content) + ']')
    print(all_content)
    print('**********Excel单元格的内容读取完毕**********')
    obj = Classify()
    pred = obj.analyse_model(all_content)
    predict = argmax(pred, axis=1) #axis = 1是取行的最大值的索引，0是列的最大值的索引
    result=[]
    analysePositive,analyseNegtive = 0,0
    for i, val in enumerate(predict):
        if (val.astype(numpy.int32) == 0):
            analysePositive+=1
            result.append("负")
        if (val.astype(numpy.int32) == 1):
            result.append("正")
            analyseNegtive += 1
    xlsc = copy(excel_file)
    shtc = xlsc.get_sheet(0)
    # (行,列,要追加的值)
    for i, val in enumerate(result):
        shtc.write(i, 1, result[i])

    result_filename = "download"+Pic_str().create_uuid()+'.xls'
    xlsc.save(os.path.join(file_dir,result_filename))  # 保存文件名


    a = range(2, 10, 1)  # 主题个数
    obj = Clustering()
    a = range(2, 6, 1)  # 主题个数
    prep1, prep2, cohers1, cohers2=obj.train_model(a,excel_path)
    obj.graph_draw(a, prep1, prep2, cohers1, cohers2)

    data = {}
    THEME = []
    with open("theme/bad1theme2.txt", mode="rt", encoding="utf-8") as f:
        res = f.read()
        print(res)
        a = res.split("\'")[1]
        THEME.append(a)
        data['THEME1'] = a
        a = res.split("\'")[3]
        THEME.append(a)
        data['THEME2'] = a
        a = res.split("\'")[5]
        THEME.append(a)
        data['THEME3'] = a
        a = res.split("\'")[7]
        THEME.append(a)
        data['THEME4'] = a
        a = res.split("\'")[9]
        THEME.append(a)
        data['THEME5'] = a
    data['analysePositive'] = analysePositive
    data['analyseNegtive'] = analyseNegtive
    data['txt1'] = "bad1theme2.doc"
    data['txt2'] = "bad2theme2.doc"
    data['totalbad'] = "totalbad.jpg"
    data['perlbad'] = "perlbad.jpg"

    print(THEME)
    print(data)
    return data

    # 储存源文件
    # f = request.files['photo']
    # if f and allowed_file(f.filename):
    #     fname = f.filename
    #     print(fname)
    #     ext = fname.rsplit('.', 1)[1]
    #     new_filename = Pic_str().create_uuid() + '.' + ext
    #     print(new_filename)
    #     f.save(os.path.join(file_dir, new_filename))
    #
    #     image_data = open(os.path.join(file_dir, '%s' % new_filename), "rb").read()
    #     response = make_response(image_data)
    #     response.headers['Content-Type'] = 	"application/vnd.ms-excel"
    #     return jsonify({"success": 0, "msg": "上传成功"})
    # else:
    #     return jsonify({"error": 1001, "msg": "上传失败"})

    # obj = Classify()
    # obj.load_model()
    # return 'Hello World!'
@app.route('/download/<string:filename>', methods=['GET'])
def download(filename):
    if request.method == "GET":
        if os.path.isfile(os.path.join('upload', filename)):
            # return send_from_directory('upload', filename, as_attachment=True)
            # return app.send_static_file(filename)
            file_dir = os.path.join(basedir, app.config['UPLOAD_FOLDER'])
            excel_path = os.path.join(file_dir, filename)  # Excel文件路径
            results = open(excel_path, 'rb').read()
            return Response(results, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            headers={"Content-Disposition": 'attachment; filename=sample.xlsx'})

        pass

@app.route('/theme')
def theme():

    # a = range(2, 10, 1)  # 主题个数
    a = range(2, 21, 1)  # 主题个数
    prep1, prep2, cohers1, cohers2=objTHeme.train_model(a,)
    obj.graph_draw(a, prep1, prep2, cohers1, cohers2)
@app.route('/emotion/<string:text>', methods=['GET'])
def emotion(text):
    happy,reliable,surprise,praise,disappointed,anger,value = obj.load_review(text)
    data = {}
    data['happy'] = happy
    data['reliable'] = reliable
    data['surprise'] = surprise
    data['praise'] = praise
    data['disappointed'] = disappointed
    data['anger'] = anger
    data['value'] = value
    # value = 50 + happy + reliable + surprise + praise + disappointed + anger
    return json.dumps(data,ensure_ascii=False)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1] in ALLOWED_EXTENSIONS

def loadData():
    # excel_path = os.path.join(os.getcwd(), 'extra.xlsx')  # Excel文件路径
    excel_path = "C:/Users/司丹/Desktop/test.xls"
    print('Excel文件的路径：' + excel_path)
    excel_file = xlrd.open_workbook(excel_path)             # 打开Excel文件
    table = excel_file.sheets()[0]                          # 通过索引打开
    print('已经打开的工作簿的名字：' + table.name)
    print('**********开始读取Excel单元格的内容**********')
    all_content = []
    for i in range(table.nrows):
        cell_value = table.cell(i, 0).value  # 获取单元格内容
        all_content.append(cell_value)
        # Logger().info('[' + ', '.join("'" + str(element) + "'" for element in row_content) + ']')
    print(all_content)
    print('**********Excel单元格的内容读取完毕**********')
    # obj = Classify()
    # obj.load_model()

if __name__ == '__main__':
    app.run(host='0.0.0.0',port=5000)


